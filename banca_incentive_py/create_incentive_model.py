from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.comments import Comment

# Create workbook
wb = Workbook()

# --- SHEET 1: INPUTS ---
ws1 = wb.active
ws1.title = "INPUTS"

inputs = [
    ("Total RMs", 500),
    ("Current GWP (Cr)", 443),
    ("Target GWP (Cr)", 600),
    ("Aspirational GWP (Cr)", 800),
    ("Current Monthly RM Productivity", 550000),
    ("Target Monthly RM Productivity", 750000),
    ("Aspirational Monthly RM Productivity", 1000000),
    ("Base Incentive %", 0.012),
    ("TL Incentive %", 0.005),
    ("BPI Weight - Topline", 0.5),
    ("BPI Weight - Activation", 0.2),
    ("BPI Weight - Product Mix", 0.15),
    ("BPI Weight - Persistency", 0.1),
    ("BPI Weight - Digital", 0.05),
]

for i, (label, value) in enumerate(inputs, start=1):
    ws1[f"A{i}"] = label
    ws1[f"B{i}"] = value

# --- SHEET 2: RM_SIMULATOR ---
ws2 = wb.create_sheet("RM_SIMULATOR")
headers = ["RM Type","Target","Achieved","% Achievement","Activation","Product Mix",
           "Persistency","Digital","BPI Score","Earnings","Quality Factor","Final Earnings"]
ws2.append(headers)

rows = [
    ["Low",750000,400000,"=C2/B2",0.5,0.4,0.6,0.3,
     "=MIN((D2*0.5)+(E2*0.2)+(F2*0.15)+(G2*0.1)+(H2*0.05),1.5)",
     "=IF(I2<0.6,C2*INPUTS!B8*0.3,IF(I2<0.8,C2*INPUTS!B8*0.6,IF(I2<1,C2*INPUTS!B8*1,IF(I2<1.2,C2*INPUTS!B8*1.5,C2*INPUTS!B8*2))))",
     0.7,"=J2*K2"],
    ["Average",750000,750000,"=C3/B3",0.8,0.7,0.8,0.6,
     "=MIN((D3*0.5)+(E3*0.2)+(F3*0.15)+(G3*0.1)+(H3*0.05),1.5)",
     "=IF(I3<0.6,C3*INPUTS!B8*0.3,IF(I3<0.8,C3*INPUTS!B8*0.6,IF(I3<1,C3*INPUTS!B8*1,IF(I3<1.2,C3*INPUTS!B8*1.5,C3*INPUTS!B8*2))))",
     1,"=J3*K3"],
    ["High",750000,1200000,"=C4/B4",1,0.9,0.9,0.8,
     "=MIN((D4*0.5)+(E4*0.2)+(F4*0.15)+(G4*0.1)+(H4*0.05),1.5)",
     "=IF(I4<0.6,C4*INPUTS!B8*0.3,IF(I4<0.8,C4*INPUTS!B8*0.6,IF(I4<1,C4*INPUTS!B8*1,IF(I4<1.2,C4*INPUTS!B8*1.5,C4*INPUTS!B8*2))))",
     1,"=J4*K4"],
]
for r in rows:
    ws2.append(r)

# --- SHEET 3: PNL_DASHBOARD ---
ws3 = wb.create_sheet("PNL_DASHBOARD")
ws3["A1"], ws3["B1"] = "Total RM Business", "=INPUTS!B6*12*INPUTS!B1"
ws3["A2"], ws3["B2"] = "Base Incentive Cost", "=B1*INPUTS!B8"
ws3["A3"], ws3["B3"] = "Weighted Multiplier", "=(0.3*0.6)+(0.5*1)+(0.2*1.8)"
ws3["A4"], ws3["B4"] = "Adjusted Incentive", "=B2*B3"
ws3["A5"], ws3["B5"] = "TL Cost", "=B1*INPUTS!B9"
ws3["A6"], ws3["B6"] = "Total Incentive Cost", "=B4+B5"
ws3["A7"], ws3["B7"] = "Incentive % of GWP", "=B6/INPUTS!B2"

# Highlight key outputs
for cell in ["B6","B7"]:
    ws3[cell].font = Font(bold=True)
    ws3[cell].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

# Add explanatory comments
ws3["B7"].comment = Comment("Shows incentive cost as % of GWP. Target range: 1–1.5%.", "Model")

# Save workbook
wb.save("Bancassurance Incentive Model.xlsx")
print("Workbook created: Bancassurance Incentive Model.xlsx")
